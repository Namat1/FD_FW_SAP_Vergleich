import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="FD/FW Liefertag Vergleich", layout="wide")
st.title("FD / FW Liefertag Vergleich")
st.caption("Prüft: Sind alle FW-Liefertage auch in FD vorhanden?")

WOCHENTAGE = {1: "Mo", 2: "Di", 3: "Mi", 4: "Do", 5: "Fr", 6: "Sa", 7: "So"}

def load_file(file):
    df = pd.read_excel(file, header=0, dtype={0: str})
    cols = df.columns.tolist()
    df = df.rename(columns={cols[0]: "SAP", cols[1]: "Name", cols[6]: "Liefertag"})
    df["SAP"] = df["SAP"].astype(str).str.strip()
    df["Liefertag"] = pd.to_numeric(df["Liefertag"], errors="coerce")
    return df[["SAP", "Name", "Liefertag"]].dropna(subset=["Liefertag"])

def tage_str(tage_set):
    sorted_tage = sorted([int(t) for t in tage_set if pd.notna(t)])
    return ", ".join([WOCHENTAGE.get(t, str(t)) for t in sorted_tage])

col1, col2 = st.columns(2)
with col1:
    fd_file = st.file_uploader("FD-Datei (.xlsx)", type=["xlsx"], key="fd")
with col2:
    fw_file = st.file_uploader("FW-Datei (.xlsx)", type=["xlsx"], key="fw")

st.divider()

sap_input = st.text_area(
    "SAP-Nummern filtern (eine pro Zeile, leer = alle)",
    height=120,
    placeholder="12823\n12920\n12923"
)

run = st.button("Vergleich starten", type="primary", disabled=(fd_file is None or fw_file is None))

if run:
    fd_df = load_file(fd_file)
    fw_df = load_file(fw_file)

    filter_sap = []
    if sap_input.strip():
        filter_sap = [s.strip() for s in sap_input.strip().splitlines() if s.strip()]

    all_sap = sorted(set(fw_df["SAP"]))
    if filter_sap:
        all_sap = [s for s in all_sap if s in filter_sap]
        not_found = [s for s in filter_sap if s not in set(fw_df["SAP"])]
        if not_found:
            st.warning(f"In FW nicht gefunden: {', '.join(not_found)}")

    rows = []
    for sap in all_sap:
        fd_rows = fd_df[fd_df["SAP"] == sap]
        fw_rows = fw_df[fw_df["SAP"] == sap]

        name = fw_rows["Name"].iloc[0] if not fw_rows.empty else (fd_rows["Name"].iloc[0] if not fd_rows.empty else "")
        fd_tage = set(fd_rows["Liefertag"].dropna().astype(int))
        fw_tage = set(fw_rows["Liefertag"].dropna().astype(int))
        fw_fehlt_in_fd = fw_tage - fd_tage

        if not fd_tage:
            status = "⛔ Nicht in FD"
        elif fw_fehlt_in_fd:
            status = "⚠️ Tage fehlen in FD"
        else:
            status = "✅ OK"

        rows.append({
            "Status": status,
            "SAP": sap,
            "Name": name,
            "FW Tage": tage_str(fw_tage) if fw_tage else "–",
            "FD Tage": tage_str(fd_tage) if fd_tage else "–",
            "FW-Tage FEHLEN in FD - bitte löschen!": tage_str(fw_fehlt_in_fd) if fw_fehlt_in_fd else "–",
            "_fw_fehlt": fw_fehlt_in_fd,
            "_kein_fd": not bool(fd_tage),
        })

    result_df = pd.DataFrame(rows)

    total = len(result_df)
    ok = (result_df["Status"] == "✅ OK").sum()
    fehlend = (result_df["Status"] == "⚠️ Tage fehlen in FD").sum()
    kein_fd = (result_df["Status"] == "⛔ Nicht in FD").sum()

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("FW-Kunden geprüft", total)
    c2.metric("✅ Vollständig in FD", ok)
    c3.metric("⚠️ Tage fehlen in FD", fehlend,
              delta=f"-{fehlend}" if fehlend else None, delta_color="inverse")
    c4.metric("⛔ Gar nicht in FD", kein_fd,
              delta=f"-{kein_fd}" if kein_fd else None, delta_color="inverse")

    st.divider()

    show_cols = ["Status", "SAP", "Name", "FW Tage", "FD Tage", "FW-Tage FEHLEN in FD - bitte löschen!"]

    filter_mode = st.radio(
        "Anzeigen:",
        ["Alle", "Nur Probleme (⚠️ + ⛔)", "Nur fehlende Tage (⚠️)", "Gar nicht in FD (⛔)"],
        horizontal=True
    )

    view_df = result_df[show_cols].copy()
    if filter_mode == "Nur Probleme (⚠️ + ⛔)":
        view_df = view_df[result_df["Status"] != "✅ OK"]
    elif filter_mode == "Nur fehlende Tage (⚠️)":
        view_df = view_df[result_df["Status"] == "⚠️ Tage fehlen in FD"]
    elif filter_mode == "Gar nicht in FD (⛔)":
        view_df = view_df[result_df["Status"] == "⛔ Nicht in FD"]

    def highlight(row):
        if row["Status"] == "⛔ Nicht in FD":
            return ["background-color: #f8d7da"] * len(row)
        if row["Status"] == "⚠️ Tage fehlen in FD":
            return ["background-color: #fff3cd"] * len(row)
        return ["background-color: #d4edda"] * len(row)

    st.dataframe(
        view_df.style.apply(highlight, axis=1),
        use_container_width=True,
        hide_index=True,
        height=520,
    )

    # ── Excel Export ──────────────────────────────────────────────────────────
    def build_excel(df_rows):
        wb = Workbook()
        ws = wb.active
        ws.title = "Vergleich"

        # Farben wie im Screenshot
        BLUE_HDR   = "2E4B8F"   # Status-Spalte Header (dunkelblau)
        RED_HDR    = "C00000"   # FW Tage Header (rot)
        GREEN_HDR  = "375623"   # FD Tage Header (dunkelgrün)
        PURPLE_HDR = "2E4B8F"   # FEHLEN-Spalte Header (blau)
        
        OK_ROW     = "E2EFDA"   # helles Grün für OK-Zeilen
        WARN_ROW   = "FFF3CD"   # gelb für fehlende Tage
        ERR_ROW    = "F8D7DA"   # rot für nicht in FD

        white_bold = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=False)
        left   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
        thin   = Side(style="thin", color="BFBFBF")
        bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers    = ["Status",    "SAP",  "Name", "FW Tage",  "FD Tage",  "FW-Tage FEHLEN in FD - bitte löschen!"]
        hdr_colors = [BLUE_HDR,   BLUE_HDR, BLUE_HDR, RED_HDR, GREEN_HDR,  PURPLE_HDR]
        col_widths = [22,          12,       38,       22,       22,         28]

        for ci, (h, color, w) in enumerate(zip(headers, hdr_colors, col_widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font  = white_bold
            cell.fill  = PatternFill("solid", start_color=color)
            cell.alignment = center
            cell.border = bdr
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.row_dimensions[1].height = 20

        data_font = Font(name="Arial", size=10)

        for ri, row in enumerate(df_rows, 2):
            if row["_kein_fd"]:
                row_fill = PatternFill("solid", start_color=ERR_ROW)
            elif row["_fw_fehlt"]:
                row_fill = PatternFill("solid", start_color=WARN_ROW)
            else:
                row_fill = PatternFill("solid", start_color=OK_ROW)

            values = [
                row["Status"], row["SAP"], row["Name"],
                row["FW Tage"], row["FD Tage"], row["FW-Tage FEHLEN in FD - bitte löschen!"]
            ]
            for ci, val in enumerate(values, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font      = data_font
                cell.fill      = row_fill
                cell.border    = bdr
                cell.alignment = left if ci == 3 else center

            ws.row_dimensions[ri].height = 16

        ws.freeze_panes  = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        # Sheet 2: Nur Probleme
        ws2 = wb.create_sheet("Probleme")
        for ci, (h, color, w) in enumerate(zip(headers, hdr_colors, col_widths), 1):
            cell = ws2.cell(row=1, column=ci, value=h)
            cell.font  = white_bold
            cell.fill  = PatternFill("solid", start_color=color)
            cell.alignment = center
            cell.border = bdr
            ws2.column_dimensions[get_column_letter(ci)].width = w
        ws2.row_dimensions[1].height = 20

        ri2 = 2
        for row in df_rows:
            if not row["_fw_fehlt"] and not row["_kein_fd"]:
                continue
            row_fill = PatternFill("solid", start_color=ERR_ROW if row["_kein_fd"] else WARN_ROW)
            values = [
                row["Status"], row["SAP"], row["Name"],
                row["FW Tage"], row["FD Tage"], row["FW-Tage FEHLEN in FD - bitte löschen!"]
            ]
            for ci, val in enumerate(values, 1):
                cell = ws2.cell(row=ri2, column=ci, value=val)
                cell.font      = data_font
                cell.fill      = row_fill
                cell.border    = bdr
                cell.alignment = left if ci == 3 else center
            ws2.row_dimensions[ri2].height = 16
            ri2 += 1

        ws2.freeze_panes   = "A2"
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    excel_buf = build_excel(rows)
    st.download_button(
        "📥 Excel herunterladen",
        data=excel_buf,
        file_name="FD_FW_Vergleich.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
